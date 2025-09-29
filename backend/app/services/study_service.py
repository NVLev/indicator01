"""
Асинхронный сервис для обработки DICOM-исследований: распаковка,
парсинг метаданных, группировка по сериям, организация файлов и
обновление статусов/результатов в БД.
"""

import os
import zipfile
import tempfile
import shutil
import time
import asyncio
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import pydicom
from pydicom.errors import InvalidDicomError
import logging

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.sql import func
from fastapi import Depends
import pandas as pd

from ..core.models import Study, StudyStatus
from ..core.db_helper import db_helper
from ..core.config import settings

logger = logging.getLogger(__name__)

# Пул потоков для синхронных операций (парсинг/IO)
thread_pool = ThreadPoolExecutor(max_workers=getattr(settings, "max_workers", 4))


class DicomProcessingError(Exception):
    """Ошибка обработки DICOM-исследования"""
    pass


class StudyStatusManager:
    """Менеджер для работы со статусами исследований"""

    # Словарь для маппинга статусов на понятные описания (рус.)
    STATUS_DESCRIPTIONS = {
        StudyStatus.UPLOADED: "Исследование загружено и ожидает обработки",
        StudyStatus.EXTRACTING: "Идет распаковка архива с DICOM-файлами",
        StudyStatus.VALIDATING: "Проверка целостности и валидности DICOM-файлов",
        StudyStatus.PROCESSING_ML: "ИИ анализирует исследование на наличие патологий",
        StudyStatus.COMPLETED: "Обработка завершена успешно",
        StudyStatus.FAILED: "Произошла ошибка при обработке",
        StudyStatus.NEEDS_REVIEW: "Требуется проверка врачом (низкая достоверность ИИ)",
    }

    @classmethod
    def get_status_description(cls, status: StudyStatus) -> str:
        """Получить человекочитаемое описание статуса"""
        return cls.STATUS_DESCRIPTIONS.get(status, "Неизвестный статус")


async def process_dicom_study(
    zip_file_path: str,
    study_id: int,
    session: AsyncSession,
    output_dir: str = "processed_studies"
) -> Dict:
    """
    Асинхронная обёртка для обработки исследования: запускает синхронную
    функцию в ThreadPool, обновляет статусы и результаты в БД.
    """
    start_time = time.time()
    processing_result: Dict = {
        "study_id": study_id,
        "processing_status": StudyStatus.COMPLETED,
        "error_message": None,
        "dicom_files": [],
        "study_metadata": {},
        "series_count": 0,
        "total_instances": 0,
        "processing_time": 0.0,
        "organized_path": None,
        "ready_for_inference": False,
    }

    try:
        # Устанавливаем статус распаковки
        await update_study_status(study_id, StudyStatus.EXTRACTING, session)

        loop = asyncio.get_event_loop()
        processing_result = await loop.run_in_executor(
            thread_pool,
            _process_dicom_study_sync,
            zip_file_path,
            study_id,
            output_dir,
            processing_result
        )

        # Записываем результаты в БД
        await update_study_results(study_id, processing_result, session)

        logger.info(
            f"Успешно обработано исследование {study_id}: "
            f"{processing_result['total_instances']} файлов"
        )

    except Exception as e:
        logger.exception(f"Ошибка обработки исследования {study_id}: {e}")
        processing_result.update({
            "processing_status": StudyStatus.FAILED,
            "error_message": str(e),
            "ready_for_inference": False
        })
        await update_study_status(study_id, StudyStatus.FAILED, session, str(e))

    finally:
        processing_result["processing_time"] = time.time() - start_time

    return processing_result


def _process_dicom_study_sync(
    zip_file_path: str,
    study_id: int,
    output_dir: str,
    processing_result: Dict
) -> Dict:
    """
    Синхронная часть обработки (будет выполняться в thread pool):
    - распаковка архива,
    - парсинг DICOM,
    - группировка по сериям,
    - организация файлов.
    """
    temp_extract_dir: Optional[str] = None
    try:
        temp_extract_dir = tempfile.mkdtemp(prefix=f"study_{study_id}_")
        logger.info(f"Извлечение исследования {study_id} во временную папку {temp_extract_dir}")

        dicom_files = extract_zip_archive(zip_file_path, temp_extract_dir)

        if not dicom_files:
            raise DicomProcessingError("В архиве не найдено DICOM-файлов")

        study_metadata, valid_files = parse_dicom_files(dicom_files)
        if not valid_files:
            raise DicomProcessingError("В архиве нет валидных DICOM-файлов")

        series_groups = group_files_by_series(valid_files)
        organized_path = organize_dicom_files(
            series_groups,
            output_dir,
            study_metadata.get("StudyInstanceUID", f"study_{study_id}")
        )

        processing_result.update({
            "dicom_files": [str(p) for p in valid_files],
            "study_metadata": study_metadata,
            "series_count": len(series_groups),
            "total_instances": len(valid_files),
            "organized_path": organized_path,
            "ready_for_inference": True,
            "processing_status": StudyStatus.COMPLETED
        })

    except Exception as e:
        processing_result.update({
            "processing_status": StudyStatus.FAILED,
            "error_message": str(e),
            "ready_for_inference": False
        })
        raise

    finally:
        if temp_extract_dir and os.path.exists(temp_extract_dir):
            try:
                shutil.rmtree(temp_extract_dir)
            except Exception as e:
                logger.warning(f"Не удалось удалить временную папку {temp_extract_dir}: {e}")

    return processing_result


async def update_study_status(
    study_id: int,
    status: StudyStatus,
    session: AsyncSession,
    error_message: Optional[str] = None
):
    """Обновление только статуса исследования и (при необходимости) сообщения об ошибке"""
    try:
        stmt = select(Study).where(Study.id == study_id)
        result = await session.execute(stmt)
        study = result.scalar_one_or_none()

        if not study:
            logger.warning(f"Исследование {study_id} не найдено")
            return

        study.processing_status = status
        if error_message:
            study.error_message = error_message

        study.updated_at = func.now()
        await session.commit()
        logger.debug(f"Статус исследования {study_id} обновлён на {status}")

    except Exception as e:
        logger.exception(f"Не удалось обновить статус исследования {study_id}: {e}")
        await session.rollback()


async def update_study_results(
    study_id: int,
    results: Dict,
    session: AsyncSession
):
    """
    Обновление записи исследования результатами обработки.
    Поддерживает разные форматы поля processing_status (enum / строка).
    """
    try:
        stmt = select(Study).where(Study.id == study_id)
        result = await session.execute(stmt)
        study = result.scalar_one_or_none()

        if not study:
            logger.warning(f"Исследование {study_id} не найдено для обновления")
            return

        metadata = results.get("study_metadata", {})

        study.probability_of_pathology = results.get("probability_of_pathology", 0.0)
        study.pathology = results.get("pathology", 0)
        study.most_dangerous_pathology_type = results.get("most_dangerous_pathology_type", "")
        study.pathology_localization_coords = results.get("pathology_localization_coords")
        study.heatmap_path = results.get("heatmap_path", "")
        study.heatmap_format = results.get("heatmap_format", "")
        study.heatmap_metadata = results.get("heatmap_metadata")
        study.study_uid = metadata.get("StudyInstanceUID", "")
        study.series_uid = metadata.get("SeriesInstanceUID", "")
        study.path_to_study = results.get("organized_path", "") or study.path_to_study

        status_raw = results.get("processing_status", StudyStatus.FAILED)
        resolved_status: StudyStatus = StudyStatus.FAILED

        if isinstance(status_raw, StudyStatus):
            resolved_status = status_raw
        else:
            try:
                resolved_status = StudyStatus(status_raw)
            except Exception:
                legacy_map = {
                    "Success": StudyStatus.COMPLETED,
                    "Failure": StudyStatus.FAILED,
                    "Uploaded": StudyStatus.UPLOADED,
                    "Extracting": StudyStatus.EXTRACTING,
                }
                resolved_status = legacy_map.get(status_raw, StudyStatus.FAILED)

        study.processing_status = resolved_status
        study.time_of_processing = results.get("processing_time", study.time_of_processing)
        study.total_instances = results.get("total_instances", study.total_instances or 0)
        study.series_count = results.get("series_count", study.series_count or 0)

        # Сохраняем метаданные как JSON (если поле есть)
        if hasattr(study, "metadata_json"):
            study.metadata_json = metadata

        study.updated_at = func.now()
        await session.commit()
        logger.info(f"Исследование {study_id} обновлено результатами обработки (status={resolved_status})")

    except Exception as e:
        logger.exception(f"Не удалось обновить результаты исследования {study_id}: {e}")
        await session.rollback()


def parse_dicom_files(file_paths: List[Path]) -> Tuple[Dict, List[Path]]:
    """Парсинг DICOM файлов и извлечение метаданных"""
    study_metadata: Dict = {}
    valid_files: List[Path] = []
    series_info: Dict[str, Dict] = {}

    for file_path in file_paths:
        try:
            ds = pydicom.dcmread(str(file_path), force=True)
            if not hasattr(ds, "StudyInstanceUID"):
                logger.warning(f"Файл {file_path} без StudyInstanceUID — пропускаем")
                continue

            if not study_metadata:
                study_metadata = extract_study_metadata(ds)

            series_uid = getattr(ds, "SeriesInstanceUID", "unknown")
            if series_uid not in series_info:
                series_info[series_uid] = {
                    "SeriesInstanceUID": series_uid,
                    "SeriesDescription": getattr(ds, "SeriesDescription", ""),
                    "SeriesNumber": getattr(ds, "SeriesNumber", ""),
                    "Modality": getattr(ds, "Modality", ""),
                    "files": []
                }

            series_info[series_uid]["files"].append(file_path)
            valid_files.append(file_path)

        except InvalidDicomError:
            logger.warning(f"Файл {file_path} не является корректным DICOM")
        except Exception as e:
            logger.warning(f"Ошибка чтения {file_path}: {e}")

    study_metadata["series_info"] = series_info
    return study_metadata, valid_files


def extract_study_metadata(ds: pydicom.Dataset) -> Dict:
    """Извлекаем полезные метаданные из первого валидного DICOM-файла"""
    metadata: Dict = {
        "StudyInstanceUID": getattr(ds, "StudyInstanceUID", ""),
        "SeriesInstanceUID": getattr(ds, "SeriesInstanceUID", ""),
        "PatientID": getattr(ds, "PatientID", ""),
        "StudyDescription": getattr(ds, "StudyDescription", ""),
        "StudyDate": getattr(ds, "StudyDate", ""),
        "StudyTime": getattr(ds, "StudyTime", ""),
        "Modality": getattr(ds, "Modality", ""),
        "Manufacturer": getattr(ds, "Manufacturer", ""),
        "ManufacturerModelName": getattr(ds, "ManufacturerModelName", ""),
    }

    if hasattr(ds, "Rows") and hasattr(ds, "Columns"):
        metadata["ImageDimensions"] = f"{ds.Rows}x{ds.Columns}"

    if hasattr(ds, "PixelSpacing"):
        metadata["PixelSpacing"] = list(ds.PixelSpacing)

    if hasattr(ds, "SliceThickness"):
        try:
            metadata["SliceThickness"] = float(ds.SliceThickness)
        except Exception:
            pass

    return metadata


def group_files_by_series(file_paths: List[Path]) -> Dict[str, List[Path]]:
    """Группируем файлы по SeriesInstanceUID (чтобы собрать серии)"""
    series_groups: Dict[str, List[Path]] = {}
    for file_path in file_paths:
        try:
            ds = pydicom.dcmread(str(file_path), stop_before_pixels=True)
            series_uid = getattr(ds, "SeriesInstanceUID", "unknown")
            series_groups.setdefault(series_uid, []).append(file_path)
        except Exception as e:
            logger.warning(f"Ошибка группировки файла {file_path}: {e}")
    return series_groups


def organize_dicom_files(series_groups: Dict[str, List[Path]], output_dir: str, study_uid: str) -> str:
    """Копируем DICOM-файлы в плоскую структуру (все файлы в корне исследования)"""

    base_dir = Path("/app/processed_studies")
    study_dir = base_dir / study_uid
    study_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"📁 Создание плоской структуры в: {study_dir}")

    total_files = 0
    file_counter = 0

    # Собираем все файлы из всех серий
    all_files = []
    for series_uid, files in series_groups.items():
        all_files.extend(files)

    # Копируем все файлы в корень директории исследования с простыми именами
    for file_path in all_files:
        try:
            # Создаем простое имя файла
            new_filename = f"image_{file_counter:05d}.dcm"
            new_path = study_dir / new_filename
            shutil.copy2(file_path, new_path)

            # Проверяем, что файл скопировался
            if new_path.exists():
                total_files += 1
                file_counter += 1
                logger.debug(f"✅ Скопирован: {file_path.name} -> {new_path.name}")
            else:
                logger.error(f"❌ Файл не скопировался: {file_path}")

        except Exception as e:
            logger.error(f"❌ Ошибка копирования {file_path}: {e}")

    logger.info(f"✅ Успешно скопировано {total_files} файлов в плоскую структуру")

    # Проверяем результат
    dcm_files = list(study_dir.glob("*.dcm"))
    logger.info(f"🔍 Проверка: в {study_dir} найдено {len(dcm_files)} .dcm файлов")

    if len(dcm_files) == 0:
        logger.error("❌ КРИТИЧЕСКАЯ ОШИБКА: Файлы не скопировались в плоскую структуру!")
        # Попробуем альтернативный метод с сохранением оригинальных имен
        return organize_with_original_names(series_groups, study_dir, study_uid)

    return str(study_dir)


def organize_with_original_names(series_groups: Dict[str, List[Path]], study_dir: Path, study_uid: str) -> str:
    """Альтернативный метод: копируем с оригинальными именами файлов"""

    logger.info("🔄 Попытка копирования с оригинальными именами")

    total_files = 0

    # Копируем все файлы, сохраняя оригинальные имена
    for series_uid, files in series_groups.items():
        for i, file_path in enumerate(files):
            try:
                # Используем оригинальное имя файла
                original_name = file_path.name
                new_path = study_dir / original_name

                # Если файл с таким именем уже существует, добавляем суффикс
                if new_path.exists():
                    new_path = study_dir / f"{i:04d}_{original_name}"

                shutil.copy2(file_path, new_path)
                total_files += 1
                logger.debug(f"✅ Скопирован с оригинальным именем: {original_name}")

            except Exception as e:
                logger.error(f"❌ Ошибка копирования {file_path}: {e}")

    logger.info(f"✅ Скопировано {total_files} файлов с оригинальными именами")
    return str(study_dir)

def extract_zip_archive(zip_path: str, extract_to: str) -> List[Path]:
    """
    Распаковка ZIP-архива с базовой фильтрацией: исключаем директории,
    скрытые файлы, и пытаемся отфильтровать не-DICOM.
    """
    extracted_files: List[Path] = []
    logger.info("Начало обработки")
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            bad_file = zip_ref.testzip()
            if bad_file:
                raise DicomProcessingError(f"Повреждённый файл в архиве: {bad_file}")


        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            file_list = [
                f for f in zip_ref.namelist()
                if not f.endswith("/") and not os.path.basename(f).startswith(".")
            ]
            logger.info(f" Найдено {len(file_list)} файлов в архиве: {file_list[:5]}...")
            if not file_list:
                raise DicomProcessingError("ZIP-архив пуст")

            os.makedirs(extract_to, exist_ok=True)
            for file_name in file_list:
                try:
                    extracted_path_str = zip_ref.extract(file_name, extract_to)
                    extracted_path = Path(extracted_path_str)
                    logger.debug(f"Checking if {extracted_path.name} is DICOM...")
                    if is_likely_dicom_file(extracted_path):
                        extracted_files.append(extracted_path)
                        logger.debug(f"✓ {extracted_path.name} identified as DICOM")
                    else:
                        logger.debug(f"✗ {extracted_path.name} not identified as DICOM")
                        try:
                            os.remove(extracted_path)
                        except Exception:
                            pass
                except Exception as e:
                    logger.warning(f"Не удалось извлечь {file_name}: {e}")

        if not extracted_files:
            raise DicomProcessingError("В архиве не найдено DICOM-файлов после фильтрации")

        return extracted_files

    except zipfile.BadZipFile:
        raise DicomProcessingError("Файл не является ZIP-архивом или повреждён")

    except PermissionError:
        raise DicomProcessingError("Отказ в доступе при распаковке архива")
    except Exception as e:
        raise DicomProcessingError(f"Ошибка распаковки архива: {e}")


def is_likely_dicom_file(file_path: Path) -> bool:
    """Эвристическая проверка, похож ли файл на DICOM (по размеру, расширению, сигнатуре)"""
    try:
        if not file_path.exists() or file_path.stat().st_size < 128:  # Reduced from 1024
            return False

        dicom_extensions = {".dcm", ".dic", ".dicom", ""}
        file_suffix = file_path.suffix.lower()

        skip_extensions = {".txt", ".log", ".xml", ".json", ".zip", ".rar", ".exe", ".dll"}
        if file_suffix in skip_extensions:
            return False

        if file_suffix not in dicom_extensions and file_suffix != "":
            pass

        with open(file_path, "rb") as f:
            header = f.read(132)

        if len(header) >= 132 and header[128:132] == b"DICM":
            return True

        if len(header) >= 4 and header[0:4] in [b"DICM", b"MEDI", b"ACR"]:
            return True


        if len(header) >= 8:
            for i in range(0, min(64, len(header) - 4), 2):
                if header[i:i + 2] in [b'\x02\x00', b'\x08\x00', b'\x10\x00', b'\x20\x00']:
                    return True

        if file_suffix == "":
            return True

        return False

    except Exception as e:
        logger.debug(f"Error checking if {file_path} is DICOM: {e}")
        # When in doubt, let pydicom decide later
        return True

class StudyService:
    """Сервис для CRUD-операций с исследованиями"""

    @classmethod
    async def create_study(
        cls,
        user_id: int,
        filename: str,
        file_path: str,
        session: AsyncSession = Depends(db_helper.session_getter)
    ) -> Study:
        """Создать новое исследование в базе"""
        try:
            study = Study(
                user_id=user_id,
                filename=filename,
                file_path=file_path,
                processing_status=StudyStatus.UPLOADED,
                created_at=func.now(),
                updated_at=func.now()
            )
            session.add(study)
            await session.commit()
            await session.refresh(study)
            logger.info(f"Создано исследование id={study.id} для user_id={user_id}")
            return study
        except Exception as e:
            await session.rollback()
            logger.exception(f"Ошибка при создании исследования: {e}")
            raise

    @classmethod
    async def get_study(
        cls,
        study_id: int,
        user_id: int,
        session: AsyncSession
    ) -> Optional[Study]:
        """Получить исследование по ID и ID пользователя (ограничение доступа)"""
        try:
            stmt = select(Study).where(
                Study.id == study_id,
                Study.user_id == user_id
            )
            result = await session.execute(stmt)
            return result.scalar_one_or_none()
        except Exception as e:
            logger.exception(f"Ошибка при получении исследования {study_id}: {e}")
            return None

    @classmethod
    async def get_user_studies(
        cls,
        user_id: int,
        session: AsyncSession,
        limit: int = 50,
        offset: int = 0
    ) -> List[Study]:
        """Получить список исследований пользователя (пагинация)"""
        try:
            stmt = (
                select(Study)
                .where(Study.user_id == user_id)
                .order_by(Study.created_at.desc())
                .limit(limit)
                .offset(offset)
            )
            result = await session.execute(stmt)
            return result.scalars().all()
        except Exception as e:
            logger.exception(f"Ошибка при получении исследований пользователя {user_id}: {e}")
            return []


#----------БЛОК ДЛЯ ДЕМО РЕЖИМА---------------------------
    @classmethod
    async def create_demo_study(
            cls,
            filename: str,
            file_path: str,
            session: AsyncSession
    ) -> Study:
        """Создать исследование для демо-режима (без привязки к пользователю)"""
        try:
            study = Study(
                # user_id=None,  # Если поле nullable=True
                user_id=1,  # Или используем фиктивный ID
                filename=filename,
                file_path=file_path,
                processing_status=StudyStatus.UPLOADED,
                created_at=func.now(),
                updated_at=func.now()
            )
            session.add(study)
            await session.commit()
            await session.refresh(study)
            logger.info(f"Создано демо-исследование id={study.id}")
            return study
        except Exception as e:
            await session.rollback()
            logger.exception(f"Ошибка при создании демо-исследования: {e}")
            raise
    @classmethod
    async def get_all_studies(cls, session: AsyncSession, limit: int = 100):
        """Получить все исследования (для демо)"""
        from sqlalchemy import select
        from app.core.models import Study

        stmt = select(Study).order_by(Study.created_at.desc()).limit(limit)
        result = await session.execute(stmt)
        return result.scalars().all()

    @classmethod
    async def get_study_by_id(cls, study_id: int, session: AsyncSession):
        """Получить исследование по ID без проверки пользователя"""
        from sqlalchemy import select
        from app.core.models import Study

        stmt = select(Study).where(Study.id == study_id)
        result = await session.execute(stmt)
        return result.scalar_one_or_none()

#-------------КОНЕЦ БЛОКА---------------------------------------------

def create_excel_report(processing_results: List[Dict[str, Any]], output_path: str) -> str:
    """Формирует улучшенный Excel-отчет в формате ТЗ с форматированием."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from datetime import datetime

    report_data = []

    for result in processing_results:
        study_metadata = result.get("study_metadata", {})

        raw_status = result.get("processing_status", "Failure")
        if str(raw_status).lower() in {"completed", "success", "ok", "обработано"}:
            status_str = "Success"
        elif str(raw_status).lower() in {"needsreview", "needs_review", "требует_проверки"}:
            status_str = "Needs Review"
        else:
            status_str = "Failure"

        pathology_type = result.get("most_dangerous_pathology_type", "")
        if not pathology_type and result.get("pathology", 0) == 1:
            pathology_type = "Патология обнаружена"

        row = {
            "path_to_study": result.get("organized_path", ""),
            "study_uid": study_metadata.get("StudyInstanceUID", ""),
            "series_uid": study_metadata.get("SeriesInstanceUID", ""),
            "probability_of_pathology": round(result.get("probability_of_pathology", 0.0), 4),
            "pathology": result.get("pathology", 0),
            "processing_status": status_str,
            "time_of_processing": round(result.get("processing_time", 0.0), 2),
            "most_dangerous_pathology_type": pathology_type,
            "pathology_localization": "",
            "generation_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        loc = result.get("pathology_localization_coords")
        if isinstance(loc, dict):
            # Извлекаем координаты в правильном порядке
            coords = [
                str(loc.get("x_min", "")),
                str(loc.get("x_max", "")),
                str(loc.get("y_min", "")),
                str(loc.get("y_max", "")),
                str(loc.get("z_min", "")),
                str(loc.get("z_max", ""))
            ]
            row["pathology_localization"] = ",".join(coords)
            print(f"Локализация добавлена: {row['pathology_localization']}")  # Для отладки
        else:
            print(f"Локализация не найдена или неверный формат: {loc}")


        report_data.append(row)


    df = pd.DataFrame(report_data)


    column_names = {
        "path_to_study": "path_to_study",
        "study_uid": "study_uid",
        "series_uid": "series_uid",
        "probability_of_pathology": "probability_of_pathology",
        "pathology": "pathology",
        "processing_status": "processing_status",
        "time_of_processing": "time_of_processing",
        "most_dangerous_pathology_type": "most_dangerous_pathology_type",
        "pathology_localization": "pathology_localization",
        "generation_date": "generation_date"
    }

    df = df.rename(columns=column_names)

    df.to_excel(output_path, index=False, engine='openpyxl')

    wb = load_workbook(output_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            if cell.column == 5:
                if cell.value == 1:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")

            elif cell.column == 4:
                if isinstance(cell.value, (int, float)) and cell.value > 0.5:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


    ws.insert_rows(1, 2)
    ws['A1'] = f"Отчет по анализу КТ исследований - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=14)

    total_studies = len(report_data)
    pathology_count = sum(1 for r in report_data if r["pathology"] == 1)
    ws[
        'A2'] = f"Всего исследований: {total_studies}, Патология обнаружена: {pathology_count}, Норма: {total_studies - pathology_count}"
    ws['A2'].font = Font(italic=True)


    wb.save(output_path)

    return output_path

